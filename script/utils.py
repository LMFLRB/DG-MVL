'''
Author: Mingfei Lu
Description: 
Date: 2022-02-16 10:17:37
LastEditTime: 2022-11-14 17:17:10
'''
import os
import sys
import torch
import numpy as np
import time
import random
import yaml
import datetime
from logging import warning
from typing import Any, List, Union
from scipy.io import savemat, loadmat
from shutil import rmtree
from easydict import EasyDict as edict
from tensorboardX import SummaryWriter
import torch.optim as optim
import torch.optim.lr_scheduler as sched
import torchvision.utils as vutils
from torch.utils.data import DataLoader, TensorDataset, Dataset, Subset
from torchvision import transforms
from typing import Optional
import pandas as pd
from collections import Counter
from math import ceil
import multiprocessing
import subprocess, webbrowser
# import matplotlib
# matplotlib.use('Agg')

from sklearn.metrics import confusion_matrix, classification_report
from sklearn import preprocessing
import seaborn as sns
import uuid
import itertools

from copy import deepcopy as copy

#####################################################################################################
# ##############################　　　　dict relative           ##################################### 
def transform_to_edict(in_dict):
    in_dict=edict(in_dict)
    for key, value in in_dict.items():
        if isinstance(value, dict):
            in_dict[key]=transform_to_edict(value)
    return in_dict  

def easydict_to_dict(ed_obj):
    if not isinstance(ed_obj, edict):
        return ed_obj
    normal_dict = dict(ed_obj)
    for key, value in normal_dict.items():
        if isinstance(value, edict):
            normal_dict[key] = easydict_to_dict(value)
    return normal_dict

def make_dict_consistent(dest_dict, source_dict):
    for key, value in source_dict.items():
        if isinstance(value, dict):
            make_dict_consistent(dest_dict[key], value)
        else:
            dest_dict[key]=value
            
def cat_listed_dict(in_list):
    def cat_dict_to_list(in_dict, out_dict, initial=True):    
        for key, value in in_dict.items():
            if isinstance(value, dict):
                if initial:
                    out_dict[key]={}                
                cat_dict_to_list(value, out_dict[key], initial)
            else:
                value = value.detach().cpu().numpy() if type(value)==torch.Tensor else np.array(value)            
                if len(value.shape)<1:
                    value = np.expand_dims(value, axis=0)
                if initial:
                    out_dict[key]=[value] 
                else:
                    out_dict[key].append(value) 
    def cat_dict_of_list_to_numpy(in_dict):
        out_dict={}
        for key, value in in_dict.items():
            if isinstance(value, dict):
                out_dict[key]=cat_dict_of_list_to_numpy(value)
            else:
                out_dict[key]=np.vstack(value)
        return out_dict
    
    out_dict={}
    for num, data_dict in enumerate(in_list):
        cat_dict_to_list(data_dict, out_dict, num==0)
    return cat_dict_of_list_to_numpy(out_dict)

def expand_dict(mydict:dict) -> dict:
    ex_dict={}
    for key, value in mydict.items():
        if isinstance(value, dict):
            ex_dict=dict(ex_dict, **expand_dict(value))
        else:
            ex_dict[key] = value
    return ex_dict
# ##############################　　　　dict relative           ##################################### 
#####################################################################################################

#####################################################################################################
# ##############################　　　　file and path           ##################################### 
def get_ckpt(tb_logger, load_version: int=-1, style: str = 'train' ):    
    log_version = tb_logger.version-1 if style == 'train' else tb_logger.version    
    version = log_version if load_version==-1 else load_version

    ckpt_path = os.path.join(tb_logger.root_dir, f"version_{version}", "checkpoints")    
    ckpt_last = get_file(ckpt_path, format='.ckpt', part_name='last' )[0]
    ckpt_best = get_file(ckpt_path, format='.ckpt', part_name='epoch')[0]
    ckpt_init = get_file(ckpt_path, format='.ckpt', part_name='init_dict')[0]
    ckpt=dict(last=ckpt_last[-1] if (len(ckpt_last)>0) else None,
              best=ckpt_best[-1] if (len(ckpt_best)>0) else None,
              init=ckpt_init[-1] if (len(ckpt_init)>0) else None,)
    return ckpt

def remove_path(file_dir):
    rmtree(file_dir)

def get_file(file_dir, part_name: str=r"mypart", format: str=r".myformat"):
    # find files in file_dir with part_name or format style
    myfiles, dirs = [], []
    try:
        for root, dirs, files in os.walk(file_dir):  
            for file in files:  
                filename, fileformat = os.path.splitext(file)
                if not format == r".myformat" and not part_name==r"mypart":
                    if  fileformat == format and part_name in filename:  # find the files with assigned format and part name
                        myfiles.append(os.path.join(root, file))  
                else:
                    if  fileformat == format or part_name in filename:  # find the files with assigned format or part name
                        myfiles.append(os.path.join(root, file))  
    except:
        warning("failed to get file with the give format or part name !")
    return myfiles, dirs if len(dirs)>0 else [file_dir]

def set_differed_filename(file_dir, constpart: str=r"mypart", format: str=r".myformat"):
    num_list = []
    num_cur = 0 
    file_list = get_file(file_dir, constpart, format)[0]
    if len(file_list)==0:
        # filename=constpart+str(num_cur=0)+format
        num_cur=0
    else:
        for file in file_list:
            start_pos = file.find(constpart)
            end_pos   = file.find(format)
            if start_pos+len(constpart)<end_pos:
                num = int(file[start_pos+len(constpart)+1:end_pos])  
            else:
                num = 0
            num_list.append(num)
        if num_cur in num_list:
            num_cur = max(num_list) + 1

    return  os.path.join(file_dir, f"{constpart}_{num_cur}{format}")  # file_dir+constpart+str(num_cur)+format

def get_subfolders(folder_path):
    subfolders = []
    if os.path.exists(folder_path):
        for file_name in os.listdir(folder_path): 
            file_path = os.path.join(folder_path, file_name)
            if os.path.isdir(file_path):
                subfolders.append(file_path)
                # subfolders.extend(get_subfolders(file_path))
    return subfolders

def get_immediate_subfolders(folder_path):
    try:
        subfolders = next(os.walk(folder_path))[1]
    except:
        subfolders = []
    subfolder_paths = [os.path.join(folder_path, subfolder) for subfolder in subfolders]
    return subfolder_paths

def get_version(path, partname="version_"):
    # subfolders=get_immediate_subfolders(path)
    subfolders=get_subfolders(path)
    if len(subfolders)==0:
        return 0
    else:
        versions=[]
        for name in subfolders:
            name = name.split("/")[-1]
            if partname in name:
                versions.append(int(name[(name.find(partname)+len(partname)):]))
        if len(versions)==0:
            max_version=0
        else:
            max_version=max(versions)+1
        return max_version

def get_version_numbers(path: str, part:str="version_", path_lever: Union[int,List]=-1) -> int:
    paths=os.path.normpath(path).split(os.sep)
    numbers=[]
    for lever in (path_lever if isinstance(path_lever, list) else [path_lever]):
        string=paths[lever]
        numbers.append(int(string[string.find(part)+len(part):]))
    return numbers[0] if len(numbers)==1 else numbers

def write_results_to_txt(file_Obj, Data_to_write, count:int=-1, mode:str='w'):
    if not hasattr(write_results_to_txt, "count"):
        write_results_to_txt.count=int(0)
    write_results_to_txt.count += 1
    write_count = count if count!=-1 else write_results_to_txt.count
    with open(file_Obj, mode) as f:
        f.write(f"write data the {write_count}-th time.\n")
        if isinstance(Data_to_write, (torch.Tensor, np.ndarray, List)):
            f.write(str(Data_to_write))
        elif isinstance(Data_to_write, dict):
            for index, (key, value) in enumerate(Data_to_write.items()):
                f.write(f"\t{key}: {value}\n")

def update_callback(writer: SummaryWriter, iteration: int, update_dict: dict):
    for key, value in expand_dict(update_dict).items():
        writer.add_scalar(
                        # tag=f"{paths[-1]}/{key}", 
                        tag=key, 
                        scalar_value=value, 
                        global_step=iteration, 
                        display_name=key)
    # paths=os.path.normpath(writer.logdir).split(os.sep)
    # writer.add_scalars(main_tag=f"{paths[-2]}-{paths[-1]}", 
    #                    tag_scalar_dict=expand_dict(update_dict), 
    #                    global_step=iteration)

def format_mean(data, latex):
    """Given a list of datapoints, return a string describing their mean and
    standard error"""
    if len(data) == 0:
        return None, None, "X"
    mean = np.mean(list(data))
    err = np.std(list(data) / np.sqrt(len(data)))
    if latex:
        return "{:.2f}\pm{:.2f}".format(mean, err)
    else:
        return "{:.2f}+-{:.2f}".format(mean, err)

def print_row(row, colwidth=14, latex=False):    
    if latex:
        sep = " & "
        end_ = "\\\\"
    else:
        sep = " "*3
        end_ = " "

    def format_val(x):
        cur_colwidth = max(len(x), colwidth)
        if np.issubdtype(type(x), np.floating):
            # x = "{:.10f}".format(x) if colwidth>10 else "{:.6f}".format(x)
            x = f"{round(x, cur_colwidth-2)}"
        return str(x).ljust(cur_colwidth)[:cur_colwidth]
    string=sep.join([format_val(x) for x in row]) + end_
    print(string)
    return string

def print_metrics(metrics, colwidth:int=14, latex:bool=True, filename:str=""):
    suffix = ('.tex' if latex else'.txt')
    sys.stdout = Tee(filename+ (suffix if suffix not in filename else ''), "w")
    df = pd.DataFrame(metrics)
    print_row(list(metrics.keys()), colwidth, latex)
    for i_cv in range(len(df)):
        print_row([df.iloc[i_cv][0]]+[f"{val:.2f}" for val in df.iloc[i_cv][1:]], 
                  colwidth,latex)

    row_str=print_row(["mean"]+[format_mean(df[key],latex) for key in list(metrics.keys())[1:]], 
                      colwidth,latex)
    sys.stdout = sys.__stdout__
    return row_str[colwidth+3:]

def print_statistics(statistics, keys, colwidth:int=14, latex:bool=True, filename:str=""):
    sys.stdout = Tee(filename, "w")
    width=max(colwidth, max([len(val) for val in list(statistics.keys())+keys]))
    print_row(keys, width, latex)
    for study_name, string in statistics.items():
        print_row([study_name]+string, width, latex)
    sys.stdout = sys.__stdout__

def write_results_to_csv(statistic, headers, index, index_name, save_path):  
    from openpyxl import load_workbook
    statistic={metric: val for metric,val in zip(headers, statistic)}
    statistic_mean={metric: val.split('\pm')[0] for metric,val in statistic.items()}
    statistic_std={metric: val.split('\pm')[1] for metric,val in statistic.items()}

    filename = os.path.join(save_path, "statistics.xlsx")
    exist = os.path.exists(filename)

    def write_row(worksheet, row_data):
        row=worksheet.max_row + 1
        for j, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row, column=j, value=cell_value)
    if not exist:
        with pd.ExcelWriter(filename) as writer:
            for data,sheet_name in zip([statistic,statistic_mean,statistic_std],
                                ['statistic','mean','std']):
                df = pd.DataFrame(data, index=[index])
                df.index.name = index_name
                df.to_excel(writer, sheet_name=sheet_name, index=index, header=True)
    else:
        workbook = load_workbook(filename)
        for data,sheet_name in zip([statistic,statistic_mean,statistic_std],
                                ['statistic','mean','std']):
            worksheet = workbook[sheet_name]
            if not exist:
                write_row(worksheet, [index_name]+headers)
            write_row(worksheet, [index]+list(data.values()))
        workbook.save(filename)

    return statistic_mean,statistic_std
    
# ##############################　　　　file and path           ##################################### 
#####################################################################################################


#####################################################################################################
# ##############################       image and plot           #####################################         
def cvtColor(colorImage, cvtType, flag=False):
    if flag:
        import cv2        
        from PIL import Image
        import numpy as np
        if isinstance(colorImage, torch.Tensor):
            img_cv2 = colorImage.cpu().numpy() if colorImage.device.type=='cuda' else colorImage.numpy()
        else:
            img_cv2 = colorImage

        flag_normalize, flag_transpose, new_shape = False, False, []
        if (img_cv2.shape[-3]==3 or img_cv2.shape[-3]==1):
            img_cv2 = img_cv2.transpose([0,2,3,1])
            flag_transpose=True

        if img_cv2.max()<1.01 and img_cv2.min()<-0.01:
            img_cv2 = np.clip((img_cv2+1)/2*255+0.5,0,255)
            flag_normalize=True
        elif img_cv2.max()<1.01 and img_cv2.min()>-0.01:
            img_cv2 = np.clip((img_cv2)*255+0.5,0,255)
            flag_normalize=True
        img_cv2 = np.stack(tuple([cv2.cvtColor(np.uint8(img_cv2[i]), cvtType) for i in range(img_cv2.shape[0])]))

        if flag_normalize:
            img_cv2=(img_cv2/255.0-0.5)*2.0
        if flag_transpose:
            img_cv2=img_cv2.transpose([0,3,1,2])
        
        
        return torch.from_numpy(img_cv2).to(colorImage.device)
    else:
        return colorImage

def image_visualize(image_data, title=""):
    from PIL import Image
    import numpy as np
    import torch

    # if isinstance(image_data, torch.Tensor):  
    #     if max(image_data) < 1.2:  
    #         image_data = image_data.mul(255).add_(0.5)
    #     image_data = image_data.clamp_(0, 255).permute(1, 2, 0).numpy().astype(np.uint8)
    
    # elif isinstance(image_data, np.ndarray):           
    #     if max(image_data) < 1.2:  
    #         image_data = image_data*255+0.5
    #     image_data = np.transpose(np.clip(image_data, 0, 255), (1, 2, 0)).astype(np.uint8)
             
    if image_data.max() < 1.2:  
        image_data = image_data*255+0.5
    image_data = image_data.clamp_(0, 255).permute(1, 2, 0).numpy() if isinstance(image_data, torch.Tensor) else np.transpose(np.clip(image_data, 0, 255), (1, 2, 0))
    
    image = Image.fromarray(image_data.astype(np.uint8))   
    image.show(title)

    return image

def image_visualize_grid(img_tsr, *args, **kwargs):
    import torch
    from torchvision.transforms import ToPILImage
    from torchvision.utils import make_grid
    from PIL import Image
    import numpy as np

    if type(img_tsr) is list:
        if len(img_tsr[0].shape) == 4:
            img_tsr = torch.cat(tuple(img_tsr), dim=0)
        elif len(img_tsr[0].shape) == 3:
            img_tsr = torch.stack(tuple(img_tsr), dim=0)        
        img_tsr = make_grid(img_tsr.cpu(), *args, **kwargs)
    elif isinstance(img_tsr, np.ndarray):     
        img_tsr = make_grid(img_tsr.cpu(), *args, **kwargs)
        
    elif isinstance(img_tsr, torch.Tensor):
        img_tsr = make_grid(img_tsr, **kwargs)
        if img_tsr.max() < 1.2:
        # Add 0.5 after unnormalizing to [0, 255] to round to nearest integer
            img_tsr = img_tsr.mul(255).add_(0.5)
        img_tsr = img_tsr.clamp_(0, 255).permute(1, 2, 0).to('cpu', torch.uint8).numpy()
    
    PILimg = ToPILImage()(img_tsr)
    # PILimg.show()

    return PILimg

def save_imgrid(img_tsr, path, *args, **kwargs):
    from torchvision.transforms import ToPILImage
    from torchvision.utils import make_grid
    PILimg = ToPILImage()(make_grid(img_tsr.cpu(), *args, **kwargs))
    PILimg.save(path)
    return PILimg

def image_save_grid(img_tsr, path, show_img = False, *args, **kwargs):
    
    import torchvision.utils as vutils      
    vutils.save_image(img_tsr.cpu().data, path, *args, **kwargs)

    if show_img:
        from PIL import Image
        Image.open(path).show()

def image_add_white_edge(target, position='up_down'):
    def image_single_add(single_image, position):
        if position=='up_down':
            single_image[:,:2,:]  = torch.ones_like(single_image[:,:2,:])
            single_image[:,-2:,:] = torch.ones_like(single_image[:,-2:,:])
        else:
            single_image[:,:,:2]  = torch.ones_like(single_image[:,:,:2])
            single_image[:,:,-2:] = torch.ones_like(single_image[:,:,-2:])
            
    size = target.shape
    if len(size)==4:
        for num in range(size[0]):
            image_single_add(target[num,::], position)
    else:
        image_single_add(target, position)

def plot_confidence_ellipse(data: np.ndarray, file_path: str):
    import matplotlib.pyplot as plt
    import uuid
    from matplotlib.patches import Ellipse

    # 计算数据的均值和协方差矩阵
    mean = np.mean(data, axis=0)
    cov = np.cov(data, rowvar=False)

    # 计算椭圆的参数
    eigvals, eigvecs = np.linalg.eig(cov)
    angle = np.degrees(np.arctan2(*eigvecs[:,0][::-1]))
    width, height = 2 * np.sqrt(eigvals) * 2

    # 绘制数据的散点图和置信椭圆
    fig, ax = plt.subplots()
    # ax.scatter(data[:,0], data[:,1], s=3, alpha=0.5)
    # ax.axis('equal')
    ell =Ellipse(xy=mean, width=width, height=height, angle=angle, edgecolor='red', lw=2, facecolor='none')
    ell.set_clip_box(ax.bbox)
    ax.add_artist(ell)
    id = uuid.uuid4().hex
    plt.show()
    plt.savefig(os.path.join(file_path, f"{id}.png"))
    plt.savefig(os.path.join(file_path, f"{id}.fig"))
    plt.savefig(os.path.join(file_path, f"{id}.pdf"))
    plt.close(fig)  # 关闭 Figure 对象

def plot_hot(data: np.ndarray, file_name: str, cmap:str="Greys"):
    import matplotlib.pyplot as plt
    # 绘制热力图
    fig, ax = plt.subplots()
    im = ax.imshow(data, cmap=cmap)

    # 添加颜色条
    cbar = ax.figure.colorbar(im, ax=ax)
    
    plt.show()
    plt.savefig(f"{file_name}.pdf")
    plt.close(fig)  # 关闭 Figure 对象

def visualize_results(save_dict, show_keys:dict={}, show_name="", path:str='', single:bool=True):    
    import matplotlib.pyplot as plt
    if len(show_keys)==0:
        show_keys = list(save_dict.keys())
    if single:
        plt.figure()
        # steps = save_dict['steps_epoch']
        for key in show_keys:
            data = save_dict[key]
            if type(data)==list:
                data = np.array(data)
            else:
                data = data.flatten()
                
            scale = (data.max()-data.min()) if 'loss' in show_name else 1.0
            min_  = data.min() if 'loss' in show_name else 0.0
            plt.plot(list(range(len(data))),(data-min_)/scale, label=key)
        plt.legend()
        plt.ion()
        plt.pause(2) 
        # plt.show(block=False)
        plt.savefig(os.path.join(path,f"{show_name}.pdf"))
        plt.close()
    else:
        # steps = save_dict['steps_epoch']
        for key in show_keys:
            plt.figure()
            data = save_dict[key]
            if type(data)==list:
                data = np.array(data)
            else:
                data = data.flatten()
            plt.plot(list(range(len(data))), data, label=key)
            plt.legend()
            plt.ion()
            plt.pause(2) 
            # plt.show(block=False)
            plt.savefig(os.path.join(path,f"{show_name}-{key}.pdf"))
            plt.close()
# ##############################　　　　image and plot           ##################################### 
######################################################################################################
 
 
#####################################################################################################
# ##############################　　　　math and basic          #####################################      
def clear_all():
    for key in list(globals().keys()):
     if (not key.startswith("__")) and (key != "key"):
         globals().pop(key) 
         print(key)
    del key

def str_repeat(str, n, **kwagrs):
    ## repeat the first m letters of the given string n times     
    mode = 'list' if kwagrs.get("mode") is None else kwagrs["mode"]
    front_len = len(str) if (kwagrs.get("m") is None  or kwagrs['m'] > len(str)) else kwagrs['m']   

    front = str[:front_len]
    result = '' if mode=='cat' else []    
    for i in range(n):
        if mode=='cat':
            result = result + front
        else:
            result.append(front)
    return result

def myformat(x, precision=4):
    if isinstance(x, float):
        return f'{x:.{precision}f}'
    else:
        return x

def is_power_of_2(n: int) -> bool:        
        """
        Given input number n, returns bool of whether n is power of number 2.
        """
        if n==0:
            return True
        return (n & (n-1)) == 0

def is_power_of(n, m):
    from torch import log, tensor, equal, floor, float

    if n <= 0 or m <= 0:
        return False
    k = log(tensor(n, dtype=float)) / log(tensor(m, dtype=float))
    return equal(k, floor(k))

def get_kernelsize(features: torch.Tensor, selected_num: int=10):
    ### estimating kernelsize with data
    from scipy.spatial.distance import pdist, squareform
    features_numpy = torch.flatten(features, 1).detach().cpu().float().numpy()
    k_features = squareform(pdist(features_numpy, 'euclidean'))
    return (np.sort(k_features, -1)[:, :selected_num]).mean((0,1)) 
  
def get_optimizer(optimizer):    
    if isinstance(optimizer, dict):
        optim_func = optimizer['name']
        optimizer_params = copy(optimizer)
        del optimizer_params['name']
        optimizer_ = lambda model: getattr(optim, 
                                       optim_func)(model.parameters(), **optimizer_params)
    else:
        optimizer_  = optimizer
    return optimizer_
    
def get_scheduler(scheduler):
    if isinstance(scheduler, dict):
        sched_func = scheduler['name']
        scheduler_params = copy(scheduler)
        del scheduler_params['name']
        scheduler_ = lambda optimizer: getattr(sched, sched_func)(optimizer,**scheduler_params)
    else:
        scheduler_ = scheduler
    return scheduler_

def get_time():
    timestamp = time.time()
    localtime = time.localtime(timestamp)
    formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", localtime)
    return formatted_time

def manual_seed_all(seed: Optional[int] = None, workers: bool = False):
    os.environ["PL_GLOBAL_SEED"] = str(seed)
    # Set the random seed for PyTorch
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    # Set the random seed for numpy
    np.random.seed(seed)
    random.seed(seed)
    os.environ["PL_SEED_WORKERS"] = f"{int(workers)}"

def make_ablation_combinations(choices:Union[List,torch.Tensor,np.ndarray])-> List:
    result = []
    for i in range(1, len(choices) + 1):
        combinations = itertools.combinations(choices, i)
        result.extend(list(combinations))
    return result

def make_weights_for_balanced_classes(dataset):
    counts = Counter()
    classes = []
    for _, y in dataset:
        y = int(y)
        counts[y] += 1
        classes.append(y)

    n_classes = len(counts)

    weight_per_class = {}
    for y in counts:
        weight_per_class[y] = 1 / (counts[y] * n_classes)

    weights = torch.zeros(len(dataset))
    for i, y in enumerate(classes):
        weights[i] = weight_per_class[int(y)]

    return weights

def classification_results(labels, preds, metrics, digits=4, output_dict=True):    
    report_ = classification_report(labels, preds, digits=digits, output_dict=output_dict)
    report = {}
    inner_metrics = copy(metrics)
    if 'accuracy' in metrics:
        report['accuracy'] = report_['accuracy']
        inner_metrics.remove('accuracy')
    report.update({key: report_['macro avg'][key] for key in inner_metrics})
    return report

class myEventLoader(Dataset):
    # to transform events file to mat file
    # requirement: tensorboard,  get_file() 
    # eventloader = EventLoader(path):  #define a EventLoader with the given path including some *.events files
    # files = eventloader.events_to_mat(file_num=0): # then call this function to transform a file of given file_nums,
    #   here file_num can be int scale or int list 
    
    def __init__(self, file_path) -> None:
        self.file_path = get_file(file_path, part_name="events.out.tfevents.")[0] 
        # Numbers = [str(num) for num in range(10)]
        # self.file_path = [file for file in file_path if os.path.splitext(file)[-1][-1] in Numbers]
        if self.file_path==[]:
            warning("There're no events files in the given path")     
        
        self.Key = []
        self.Data= []
        self.EA  = None
        
    def readFile(self, **kwargs):
        from tensorboard.backend.event_processing.event_accumulator import EventAccumulator
        file_path = self.file_path[0] if kwargs.get('file_path') is None else kwargs['file_path']
        ea = EventAccumulator(file_path)
        ea.Reload()
        self.EA = ea.scalars
        self.Key = [key for key in ea.scalars.Keys()]
        self.Data = {key: ea.scalars.Items(key) for key in self.Key}

    def get_item(self, index_key=-1, index_item=-1) -> Any:
        # index_key:  key   index to transform, self.Key[index_key]
            # default sets to -1 if transforming all the Keys in data
        # index_item: items index to transform, self.data[key][index_item]
            # default sets to -1 if transforming all the items of given Key in data
        if index_key!=-1:
            index_key = [index_key] if isinstance(index_key, int) else index_key        
            Key  = [self.Key[key] for key in index_key if ( key <len(self.Key)  ) ]   
        else:
            Key  = self.Key
            
        result = {}
        for key in Key:
            if index_item != -1:
                index_out = [index_item] if isinstance(index_item, int) else index_item
            else:
                len_item = len(self.Data[key])
                index_out = [index for index in range(len_item)]
            items = [self.Data[key][index].value for index in index_out]
            key = key.replace('-','_').replace('/','_')
            if len(items)>1:    
                result[key] = list(set(np.array(items).squeeze().astype(np.int32).tolist())) if key=='epoch' else items
            elif len(items)==1:  
                result[key] = list(int([0])) if key=='epoch' else items
            else:
                result[key] = []
        return result

    def events_to_mat(self, file_num=-1, write_file=True):    
        # call this function to convert the events files of given number in path to *.mat
        from scipy.io import savemat
        if file_num==-1:
            file_list = [num for num in range(len(self.file_path))]
        else:
            file_list = [file_num] if isinstance(file_num, int) else file_num
        mat_files = []
        for num in file_list:     
            mat_file = self.file_path[num]+".mat" 
            if os.path.exists(mat_file):
                mat_files.append(mat_file)
            else:
                if num>=len(self.file_path): 
                    continue    
                self.readFile(file_path=self.file_path[num])
                results = self.get_item(index_item = -1, index_key = -1)
                mat_files.append(mat_file)
                if results=={}:
                    continue
                else:
                    if write_file:
                        savemat(mat_files[num], results)
        
        return mat_files
    
    def events_to_yaml(self, file_num=-1):   
        from scipy.io import loadmat
        import yaml
        if file_num==-1:
            file_list = [num for num in range(len(self.file_path))]
        else:
            file_list = [file_num] if isinstance(file_num, int) else file_num
        for num in file_list:     
            eventfile = self.file_path[num]    
            try:
                data=loadmat(eventfile+".mat")
                if not data=={}:
                    with open(eventfile+".yaml", 'w') as file:
                        yaml.dump(data, file) 
            except:
                pass

class myTensorboard():
    def __init__(self, log_dir, port:int=6006):
        self.log_dir = log_dir
        self.port = port      
        self.tb_started = multiprocessing.Event()  
    
    def start_tensorboard(self):
        # 构造TensorBoard命令行参数
        cmd = ["tensorboard", "--logdir", self.log_dir, "--port", str(self.port)]
        # 启动TensorBoard进程
        tb_process = subprocess.Popen(cmd)      
        # 构造TensorBoard URL地址
        tb_url = "http://localhost:{}/".format(self.port)
        # 自动打开TensorBoard页面
        webbrowser.open(tb_url)
        
        # # 分离进程，使其在父进程退出后仍能继续运行
        # tb_process.detach()
        # 将子进程设置为新的进程组，并将其分离到新的会话中
        # os.setsid()
        
        # 设置事件，通知主进程TensorBoard进程已启动完成
        self.tb_started.set()
        return tb_process

    def open_tensorboard(self, run_version:int=1):
        self.active_processes, active_ports, exist = self.get_active_tensorboard()
        start=ready=time.time()
        if exist:      
            self.tb_process=self.active_processes[0]
            self.port = active_ports[0]
        else:      
            self.tb_process = multiprocessing.Process(target=self.start_tensorboard,) 
                                                #  args=(self.log_dir, self.port, self.tb_started))
            self.tb_process.port = self.port = (max(active_ports)+1) if (6001+run_version) in active_ports else (6001+run_version)
            self.tb_process.start()
            
            # 等待TensorBoard进程启动完成
            self.tb_started.wait()   
            ready=time.time() 
            
            # 等待TensorBoard进程结束
            self.tb_process.join()

        return (start, ready)
    
    def get_active_tensorboard(self):
        # 遍历进程列表，查找并关闭活动的TensorBoard进程
        import psutil
        # 获取tensorboard进程然后关闭进程
        Process, Port , exist = [], [], False
        for process in psutil.process_iter():
            try:
                if "tensorboard" in process.name().lower():
                    try:
                        # 获取TensorBoard进程的端口号   
                        port = int(process.cmdline()[process.cmdline().index("--port")+1])    
                        # 获取TensorBoard进程的logdir   
                        logdir = process.cmdline()[process.cmdline().index("--logdir")+1] 
                        if logdir == self.log_dir:
                            process.port = port
                            Process = [process]
                            Port=[port]
                            exist = True
                            break
                        else:
                            Process.append(process)
                            Port.append(port)
                    except:
                        pass

            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        return Process, Port, exist

    def close_tensorboard(self):
        self.tb_process.terminate()
        # # 在需要关闭TensorBoard进程时，使用PID来终止进程
        # subprocess.call(["kill", str(self.tb_process.pid)])

    def close_existing_tensorboard(self):
        # 遍历进程列表，查找并关闭活动的TensorBoard进程
        import psutil
        # 获取tensorboard进程然后关闭进程
        for process in self.active_processes:
            try:
                # 终止TensorBoard进程
                process.terminate()
                print(f"TensorBoard process with PID-{process.pid} @port-{process.port} has been terminated.")
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass

class Tee():
    def __init__(self, fname, mode="a"):
        super(Tee, self).__init__()
        os.makedirs(os.path.dirname(fname), exist_ok=True)
        self.stdout = sys.__stdout__
        self.file = open(fname, mode)

    def write(self, message):
        self.stdout.write(message)
        self.file.write(message)
        self.flush()

    def flush(self):
        self.stdout.flush()
        self.file.flush()

class EarlyStop(torch.nn.Module):
    def __init__(self, 
                 patience:int=10, 
                 divation:float=1.e-3, 
                 small_good:bool=None,
                 direction:str=None,
                 silent:bool=True,
                 limit:float=None,
                 **kwargs):
        super(EarlyStop, self).__init__()
        if small_good is not None and direction is None:
            self.small_good = small_good
            self.direction = "minimize" if small_good else "maximize"
        elif small_good is None and direction is not None:
            self.direction = direction if small_good else "maximize"
            self.small_good = True if direction=="minimize" else False
        elif small_good is not None and direction is not None:
            if small_good == (direction=="minimize"):
                self.small_good = small_good
                self.direction = direction
            else:
                ValueError("please make sure input the right direction")
        else:
            Warning("please input 'small_good' or 'direction' for setting")

            
        self.limit = limit
        self.patience = patience
        self.divation = divation
        self.silent = silent
        self.errors=torch.tensor(1.0).repeat(patience+1)
        self.check_value=torch.tensor(0.0)
        self.count=0
        self.best_value = 1.0 if self.small_good else 0.0
        self.no_improve_count = 0


    def forward(self, check_value, eps=1.0e-8):            
        if not isinstance(check_value, torch.Tensor):
            check_value = torch.tensor(check_value)
        if check_value.dtype != self.check_value.dtype:
            self.errors=self.errors.to(check_value.dtype)
            self.check_value=self.check_value.to(check_value.dtype)

        self.count += 1
        self.errors[1:]=self.errors[:-1].clone()
        self.errors[0] =(check_value-self.check_value)
        if (self.best_value>check_value and self.small_good) or (self.best_value<check_value and not self.small_good):
            self.best_value=check_value
            self.no_improve_count=self.count

        varying_slow=(self.errors.abs()<self.divation*self.check_value.abs()).all().item()
        no_improving=(self.count-self.no_improve_count)>=self.patience
        aproach_limit=((check_value<=self.limit+eps) if self.small_good else (check_value>=self.limit-eps)) if self.limit is not None else False
        self.flag = aproach_limit or ((no_improving or varying_slow) if self.count>self.patience else False)
        self.check_value=check_value
        
        if self.flag:
            if aproach_limit:
                self.info="stopped with optimization limit"
            elif no_improving:
                self.info=f"stopped with no improvement for {self.patience} runs"
            else:
                self.info=f"stopped with too slowly variance for {self.patience} runs"                
        else:
            self.info="is not earlystopped"

        if not self.silent:
            print(f"procedure of check_value {self.info}")

        return self.flag

#####################################################################################################
# ##############################　　special functions for task   #################################### 
def make_configs_consistent(Data, experiment, model):
    experiment.enable_sample_weight = Data.enable_sample_weight
    if experiment.max_epochs*len(Data.train_dataloader())<experiment.min_iters:
        experiment.max_epochs = ceil(experiment.min_iters/len(Data.train_dataloader()))
    model.cuda = experiment.cuda
    model.feature_dims = Data.feature_dims
    model.n_view = int(Data.n_view)
    model.n_class = int(Data.n_class)
    model.autoencoder.n_class = int(Data.n_class)
    model.predictor.n_class = int(Data.n_class)

def save_hyperparameters(parameters, file):
    os.makedirs(os.path.dirname(file), exist_ok=True)
    params = easydict_to_dict(parameters) if isinstance(parameters, edict) else parameters
    with open(file, 'w') as f:
        yaml.safe_dump(params, f, default_flow_style=False)

def set_path(config, log_name=None, use_format_time=True):    
    workspace = "/opt/data/private/lmf/" if config.experiment.ssh else "G://"
    if log_name is None:
        log_name="-".join([f'Fuser={config.model.fuse_method}',
                        f'Common={config.model.common_method}',
                        f"Loss={'+'.join(sorted(config.loss.weights.keys()))}",
                        F"Div_uni={config.loss.div_uni}",
                        f"Metric={config.experiment.basic_metric}",
                        f'Optim={config.optimizer.name}',
                        ]) 
    root_dir = os.path.join((workspace+"logs") if config.experiment.ssh else (workspace+"Logs"),
                            config.experiment.log_dir, config.experiment.mode)
    formatted_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")    
    log_dir = os.path.join(root_dir, log_name)
    if use_format_time:
        log_dir = os.path.join(log_dir, formatted_time,)

    
    os.makedirs(log_dir, exist_ok=True)
    return workspace, root_dir, log_dir

def visualize_synthetic(to_visualize, path:str=''):    
    import matplotlib.pyplot as plt
    # visualize the groundtruth
    color_cycle = plt.rcParams['axes.prop_cycle'].by_key()['color']
    X = to_visualize.T
    Ys= [to_visualize.common_orig] + [uni for uni in to_visualize.uniques_orig] + [to_visualize.noise]
    Legs=["common"] + [f"unique_{i}" for i in range(len(to_visualize.uniques_orig))] + ["noise"]
    h_o=plt.figure("ground_truth")
    for i,Y in enumerate(Ys):
        plt.plot(X, Y, linewidth=2, color=color_cycle[i])
    plt.xlabel('t')
    plt.ylabel('x')
    plt.legend(Legs)
    plt.savefig(os.path.join(path, "ground_truth.pdf"))
    plt.ion()
    
    Ys= [to_visualize.common_trained] + [uni for uni in to_visualize.uniques_trained]
    h_t=plt.figure("trained_result")
    for i,Y in enumerate(Ys):
        plt.plot(X, Y, linewidth=2, color=color_cycle[i])
    plt.xlabel('t')
    plt.ylabel('x')
    plt.legend(Legs[:-1])
    plt.savefig(os.path.join(path, "trained_result.pdf"))

    plt.ion()
    plt.pause(5)

    plt.close(h_o)    
    plt.close(h_t)

def visualize_embeddings(latents, labels, path=""):
    from sklearn.manifold import TSNE
    import matplotlib.pyplot as plt
    # 创建图像
    fig, ax = plt.subplots(figsize=(12, 8))

    # 定义marker类型
    markers = ['o', 's', '^', 'D', 'v', 'p']

    # 遍历每一行数据,绘制散点图
    embedings=[]
    latents_=[]
    for i,latent in enumerate(latents):
        # Perform s-tne dimensionality reduction
        if isinstance(latent, torch.Tensor):
            latent=latent.detach().cpu().numpy()
        if isinstance(latent, list):
            if isinstance(latent[0], torch.Tensor):
                latent=torch.stack(latent).detach().cpu().numpy()

        tsne = TSNE(n_components=2)
        embedding_tsne = tsne.fit_transform(latent)
        embedings.append(embedding_tsne)
        latents_.append(latent)
        ax.scatter(embedding_tsne[:,0], embedding_tsne[:,1], 
                   c=labels, cmap='viridis', marker=markers[i], s=50, alpha=0.7)
    # 添加图例
    legend_elements = [plt.Line2D([0], [0], marker=markers[i], color='w', label=f'View {i+1}', markerfacecolor='k', markersize=10) 
                       for i in range(len(latents))]
    ax.legend(handles=legend_elements, loc='upper left')
    ax.set_xlabel('X')
    ax.set_ylabel('Y')   
    plt.title("S-TNE Visualization of Embeddings")
    plt.legend()
    plt.savefig(path+".pdf", bbox_inches='tight', dpi=600)
    plt.show()
    plt.close()
    savemat(path+".mat", dict(embed=latents_,
                              embedTSNE=embedings,
                              labels=labels))